import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Carga de datos
def load_csv(path: str, encoding: str = 'latin-1') -> pd.DataFrame:
    return pd.read_csv(path, encoding=encoding)

def check_required_columns(df: pd.DataFrame, required_columns:list):
    missing_columns = [col for col in required_columns if col not in df.columns]
    return {
        "test_id": "CT00",
        "description": "Validar presencia de columnas obligarias",
        "passed": len(missing_columns) == 0,
        "issues_count": len(missing_columns),
        "details": f"Faltan columnas oblogarias: {', '.join(missing_columns)}" if missing_columns else ""
    }

# Validación de datos nulos
def check_nulls(df: pd.DataFrame): # Validar ausencia de valores nulos
    null_counts = df.isnull().any(axis=1).sum() # Contar filas con al menos un valor nulo, axis=1 para filas
    return {
        "test_id": "CT01",
        "description": "Validar ausencia de valores nulos",
        "passed": null_counts == 0, # Pasa si no hay valores nulos
        "issues_count": int(null_counts),
        "details": f"{null_counts} filas con valores nulos" if null_counts > 0 else ""
    }

# Validación de tipos numéricos
def check_numeric_fields(df: pd.DataFrame):
    precio = pd.to_numeric(df["precio"], errors='coerce')
    total = pd.to_numeric(df["total_venta"], errors='coerce')
    cantidad_vendida = pd.to_numeric(df["cantidad_vendida"], errors='coerce')
    non_numeric_count = precio.isnull().sum() + total.isnull().sum() + cantidad_vendida.isnull().sum()
    return {
        "test_id": "CT02",
        "description": "Validar tipos numericos en columnas (precio, total_venta y cantidad_vendida)",
        "passed": non_numeric_count == 0,
        "issues_count": int(non_numeric_count),
        "details": f"{non_numeric_count} valores no numericos encontrados" if non_numeric_count > 0 else ""
    }

# Validación de consistencia entre precio, cantidad y total
def check_price_quantity_total(df: pd.DataFrame):
    precio = pd.to_numeric(df["precio"], errors='coerce')
    total = pd.to_numeric(df["total_venta"], errors='coerce')
    cantidad_vendida = pd.to_numeric(df["cantidad_vendida"], errors='coerce')

    calculated = precio * cantidad_vendida
    diff = np.abs(calculated - total)

    mismatches = diff > 0
    mismatch_count = mismatches.sum()
    return {
        "test_id": "CT03",
        "description": "Validar que precio * cantidad_vendida == total_venta",
        "passed": mismatch_count == 0,
        "issues_count": int(mismatch_count),
        "details": f"{mismatch_count} inconsistencias de cálculo en total_venta" if mismatch_count > 0 else ""
    }

# Validación de fechas válidas
def check_valid_dates(df: pd.DataFrame):
    fechas = pd.to_datetime(df["fecha_venta"], errors='coerce')
    invalid_date_count = fechas.isnull().sum()
    return {
        "test_id": "CT04",
        "description": "Validar formato y validez de (fechas de ventas)",
        "passed": invalid_date_count == 0,
        "issues_count": int(invalid_date_count),
        "details": f"{invalid_date_count} fechas no validas encontradas" if invalid_date_count > 0 else ""
    }

# Validación de cantidades positivas
def check_positive_quantity(df):
    invalid = df["cantidad_vendida"] <= 0
    count = invalid.sum()
    return {
        "test_id": "CT05",
        "description": "Validar que columna (cantidad_vendida) sea mayor a cero",
        "passed": count == 0,
        "issues_count": int(count),
        "details": f"{count} registros con cantidad_vendida no positiva" if count > 0 else ""
    }

# Validación de categorías permitidas
def check_allowed_categories(df):
    allowed = ["Electrónica", "Oficina", "Accesorios", "Fotografía", "Computación", "Audio"]
    # Discriminar mayúsculas y minúsculas
    invalid = ~df["categoria"].str.lower().isin([cat.lower() for cat in allowed])
    invalid_count = invalid.sum()
    return {
        "test_id": "CT06",
        "description": "Validar valores permitidos en la columna (categoria)",
        "passed": invalid_count == 0,
        "issues_count": int(invalid_count),
        "details": f"{invalid_count} valores no permitidos encontrados en categoria" if invalid_count > 0 else ""
    }

# Ejecutar todas las validaciones
def run_validations(df, validations):
    results = []
    for validation in validations:
        result = validation(df)

        results.append({
            "ID caso de prueba": result["test_id"],
            "Descripción del Caso": result["description"],
            "Resultado": "Aprobado" if result["passed"] else "Fallido",
            "Descripción del Resultado": result["details"] if not result["passed"] else "",
            "Número de Incidencias Detectadas": result["issues_count"],
            "Observaciones": generate_observations(result)
        })
    return results

def generate_observations(result):
    if result["passed"]:
        return "Validación exitosa. No se detectaron problemas."
    if result["issues_count"] > 0:
        return(
            f"Se detectaron {result['issues_count']} incidencias."
            "Revisión y corrección de datos requerida antes de continuar."
        )
    return "Validación fallida. Revisión de datos necesaria."

# Convertir resultados a DataFrame
def results_to_dataframe(results):
    return pd.DataFrame(results)

# Exportar resultados a Excel con formato de tabla
def export_results_to_excel(df: pd.DataFrame, output_path: str):
    sheet_name = "Data_Quality_Report"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{chr(64 + max_col)}{max_row}"
    table = Table(displayName="DataQualityResults", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", 
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=True,
                           showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    # Autoajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(output_path)